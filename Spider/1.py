import re
import time
import sys
import requests
from lxml import etree
import openpyxl
from openpyxl import load_workbook, Workbook

id = 1
# wb = load_workbook('./四川省考研学校收集表.xlsx')
# ws = wb['院校基本信息']

def newfile():
    wb = Workbook()
    ws = wb.active
    ws.title = "院校基本信息"
    titlelist = ['序号', '院校名称', '所在地', '院校隶属', '研究生院', '自划线院校', '院校简介', '周边环境', '院校官网地址', '研究生院官网地址']
    ws.append(titlelist)
    return wb

def get_xpath(url):
    try:
        response = requests.get(url)
        return etree.HTML(response.text)
    except Exception:
        print(url, '该页面没有相应！')
        return ''

def get_basic_facts(id,item,wb):
    dataList = [id]
    ws = wb.active
    bing = "https://cn.bing.com/search?q="
    # baidu = "https://www.baidu.com/s?ie=utf-8&f=8&rsv_bp=1&rsv_idx=1&tn=baidu&wd="
    name = item.xpath('./td[1]/a/text()')
    name = re.sub(r'\s+', "", name[0])
    surl = item.xpath('./td[1]/a/@href')
    address = item.xpath('./td[2]/text()')
    belong = item.xpath('./td[3]/text()')
    research = item.xpath('./td[4]/i/text()')
    examself = item.xpath('./td[5]/i/text()')
    surlc = 'https://yz.chsi.com.cn' + surl[0]
    dataList.append(name)
    dataList.append(address[0])
    dataList.append(belong[0])
    if len(research) != 0:
        research = '是'
    else:
        research = '否'
    dataList.append(research)
    if len(examself) != 0:
        examself = '是'
    else:
        examself = '否'
    dataList.append(examself)
    html = get_xpath(surlc)
    schoolLink = html.xpath('//ul[@class="yxk-link-list clearfix"]')
    curl = schoolLink[0].xpath('./li[1]/a/@href')
    curlc = 'https://yz.chsi.com.cn' + curl[0]
    condition = get_xpath(curlc).xpath('//div[@class="container"]')
    content = condition[0].xpath('string(./div[4])')
    env = condition[0].xpath('string(./div[6])')
    env = re.sub(r'\s+', "", env)
    content = re.sub(r'\s+', "", content)
    dataList.append(content)
    dataList.append(env)
    # dataList.append(curlc)
    for i in range(0,5):
        response = requests.get(bing + name)
        html_bing = etree.HTML(response.text)
        bing_school = html_bing.xpath('//li[@class="b_algo"][1]/h2/a/@href')
        if bing_school:
            break
        else:
            # print("error")
            continue
    # baidu_school = html_bing.xpath('//*[@id="1"]/h3/a[1]/em/text()')
    for i in range(0, 5):
        response1 = requests.get(bing + name + str("研究生院"))
        html_bing1 = etree.HTML(response1.text)
        bing_school_research = html_bing1.xpath('//li[@class="b_algo"][1]/h2/a/@href')
        if bing_school_research:
            break
        else:
            # print("error")
            continue
    # baidu_school = html_bing.xpath('//*[@id="1"]/h3/a[1]/em/text()')

    print(name)
    print(bing + name)
    print(bing_school)
    print(bing_school_research)
    try:
        dataList.append(bing_school[0])
    except:
        dataList.append('')
    try:
        dataList.append(bing_school_research[0])
    except:
        dataList.append('')
    print(dataList)


    ws.append(dataList)
    wb.save("院校基本信息.xlsx")


def all_data(url,wb):
    global id
    collegeList = get_xpath(url).xpath('//div[@class="yxk-table"]')
    for item in collegeList:
        dataList = item.xpath('./table/tbody/tr')
    for item in dataList:
        get_basic_facts(id,item,wb)
        id = id+1



def main():
    wb = newfile()
    url = "https://yz.chsi.com.cn/sch/search.do?ssdm=51&start="
    for i in range(0,2):
        all_data(url+str(i*20),wb)

if __name__ =="__main__":
    main()